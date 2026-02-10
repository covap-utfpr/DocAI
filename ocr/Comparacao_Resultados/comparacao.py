import os, sys
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from Modules.functions import extrair_tokens, granularidade, comparar_consumindo, calcular_metricas, rotular_nome

# ---------------------------------------------
# MAIN
# ---------------------------------------------
def main():                                                                         # Função principal
    if len(sys.argv) < 3:                                                           # Verifica se possui pelo menos 2 argumentos (base + 1 cupom)
        print("Uso: python3 compare.py base.txt cupom1.txt cupom2.txt ...")         # Instruções de uso
        print("Aceita arquivos .txt (OCR) e .json")                                 # Instruções de uso
        return                                                                      # Encerra a função 

    base_path = sys.argv[1]                                                         # Caminho do arquivo base    
    base_nome = os.path.splitext(os.path.basename(base_path))[0]                    # Nome da base sem extensão e caminho
    base = extrair_tokens(base_path)                                                # Extrai tokens da base (função)
    print(f"Base '{base_nome}' carregada: {len(base)} tokens\n")                    # Mensagem de status
    cupons_paths = sys.argv[2:]                                                     # Caminhos dos arquivos de cupons a serem comparados          
    nomes_cupons = [rotular_nome(p) for p in cupons_paths]                          # Gerar rótulos para cupons baseado na extensão de arquivo
    df_resultado = pd.DataFrame({"ITEM_BASE": base})                                # DataFrame para resultados (booleanos) 
    df_conteudo = pd.DataFrame({"ITEM_BASE": base})                                 # DataFrame para conteúdos encontrados 
    metricas_total = {}                                                             # Dicionário para armazenar métricas totais por cupom

    # Criar pasta para resultados e subpastas
    root_dir = "results"
    os.makedirs(root_dir, exist_ok=True)

    output_dir = os.path.join(root_dir, f"{base_nome}_results")         # Nome da pasta de resultados da detecção dentro de "results"    
    os.makedirs(output_dir, exist_ok=True)                              # Criação da pasta de resultados

    # ---------------------------------------------
    # Loop pelos cupons
    # ---------------------------------------------
    cupons_tokens = {}                                                  # Dicionário para armazenar tokens de cada cupom
    for path, nome in zip(cupons_paths, nomes_cupons):                  # Para cada caminho e nome de cupom em zip (lista paralela)
        cupom = extrair_tokens(path)                                    # Extrai tokens do cupom (função)
        cupons_tokens[nome] = cupom                                     # Armazena os tokens no dicionário
        print(f"Cupom '{nome}' carregado: {len(cupom)} tokens")         # Mensagem de status 
        contem, conteudo = comparar_consumindo(base, cupom)             # Compara base e cupom (função)
        df_resultado[nome] = contem                                     # DataFrame de resultados (booleanos)
        df_conteudo[nome] = conteudo                                    # DataFrame de conteúdos encontrados
        metricas_total[nome] = calcular_metricas(base, cupom, contem)   # Métricas do cupom (função)

    # ---------------------------------------------
    # Criar XLSX
    # ---------------------------------------------
    wb = Workbook()                                                                         # Cria novo workbook do Excel (xlsx) 
    fill_ok = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")      # Se o conteúdo foi encontrado pinte de verde
    fill_err = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")     # Se o conteúdo não foi encontrado pinte de vermelho
    fill_empate = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Se houve empate pinte de amarelo

    # ABA 1 – RESULTADO
    ws1 = wb.active                                                                 # Primeira aba  
    ws1.title = "Resultado"                                                         # Nomeada de Resultado
    ws1.append([f"BASE: {base_nome}"] + [f"COMP: {n}" for n in nomes_cupons])       # Cabeçalho com nomes
    ws1.append(list(df_resultado.columns))                                          # Cabeçalho com colunas

    for i in range(len(df_resultado)):                                              # Para cada linha do DataFrame de resultados
        for j, col in enumerate(df_resultado.columns, start=1):                     # Para cada coluna 
            cell = ws1.cell(row=i + 3, column=j)                                    # Célula recebe valor de linha e coluna  
            if col == "ITEM_BASE":                                                  # Se a coluna for ITEM_BASE 
                cell.value = df_resultado.iloc[i, 0]                                # Apenas copie o valor 
                continue                                                            # Vá para a próxima iteração 
            status = df_resultado.iloc[i, j-1]                                      # Obtenha o valor booleano  
            cell.value = "CONTÉM" if status else "NÃO CONTÉM"                       # Defina o valor da célula como CONTÉM ou NÃO CONTÉM
            cell.fill = fill_ok if status else fill_err                             # Pinte a célula de acordo com o status 

    # ABA 2 – CONTEÚDO  
    ws2 = wb.create_sheet("Conteudo")                                               # Segunda aba 

    ws2.append([f"BASE: {base_nome}"] + [f"COMP: {n}" for n in nomes_cupons])       # Cabeçalho com nomes
    ws2.append(list(df_conteudo.columns))                                           # Cabeçalho com colunas
    for i in range(len(df_conteudo)):                                               # Para cada linha do DataFrame de conteúdos
        for j, col in enumerate(df_conteudo.columns, start=1):                      # Para cada coluna
            cell = ws2.cell(row=i + 3, column=j)                                    # Célula recebe valor de linha e coluna  
            value = df_conteudo.iloc[i, j-1]                                        # Valor daa célula
            cell.value = value                                                      # Defina o valor da célulaa
            if col == "ITEM_BASE":                                                  # Se a coluna for ITEM_BASE
                continue                                                            # Vá para a próxima iteração
            status = df_resultado.iloc[i, j-1]                                      # Obtenha o valor booleano correspondente
            cell.fill = fill_ok if status else fill_err                             # Pinte a célula de acordo com o status

    # ABA 3 – MÉTRICAS
    ws3 = wb.create_sheet("Metricas")                                               # Terceira aba
    cab = ["CUPOM", "TP", "FP", "FN", "TN", "Acuracia", "Precisao", "Recall", "F1"] # Cabeçalho das métricas
    ws3.append(cab)                                                                 # Adiciona o cabeçalho
    for nome, m in metricas_total.items():                                          # Para cada cupom e suas métricas
        ws3.append([                                                                # Adiciona uma linha com as métricas
            nome, m["TP"], m["FP"], m["FN"], m["TN"],
            m["Acuracia"], m["Precisao"], m["Recall"], m["F1"]
        ])

    # ABA 4 – CONSENSUS
    ws4 = wb.create_sheet("Consensus")                                  # Quarta Aba 
    ws4.append(["ITEM_BASE", "ACERTOS", "ERROS", "CONSENSUS"])          # Cabeçalho da aba
    for i, item in enumerate(df_resultado["ITEM_BASE"]):                # Para cada item na coluna ITEM_BASE
        resultados = df_resultado.iloc[i, 1:].tolist()                  # Resultados dos cupons (exclui ITEM_BASE) 
        acertos = sum(resultados)                                       # Conta acertos (True = 1)
        erros = len(resultados) - acertos                               # Conta erros (False = 0) 
        if acertos > erros:                                             # Se acertos são maioria 
            consenso = "OK"                                             # Define consenso como OK  
            fill = fill_ok                                              # Pinta de verde
        elif erros > acertos:                                           # Se erros são maioria 
            consenso = "ERR"                                            # Define consenso como ERR 
            fill = fill_err                                             # Pinta de vermelho
        else:                                                           # Se empate
            consenso = "EMPATE"                                         # Define consenso como EMPATE
            fill = fill_empate                                          # Pinta de amarelo
        ws4.append([item, acertos, erros, consenso])                    # Adiciona a linha na planilhaa
        ws4.cell(row=ws4.max_row, column=4).fill = fill                 # Pinta a célula de consenso
    
    # ABA 5 – GRÁFICO LONGO DE ACERTOS/ERROS + MATRIZ DE GRANULARIDADE   
    if len(base) > 0 and len(nomes_cupons) > 0:                         # Verificação antes de gerar o gráfico
        ws5 = wb.create_sheet("Grafico_Granularidade")                  # Quinta aba para gráfico        
        for i, nome in enumerate(nomes_cupons):                         # Para cada cupom analisado
            ws5.append([nome] + df_resultado[nome].astype(int).tolist())# Adiciona os dados do cupom (1 para True, 0 para false)        
        # -----------------------------
        # Criar matriz de granularidade
        # -----------------------------        
        df_gran = pd.DataFrame(index=nomes_cupons, columns=nomes_cupons, dtype=float)                       # DataFrame para armazenar a matriz de granularidade
        for i, n1 in enumerate(nomes_cupons):                                                               # Para cada cupom n1
            for j, n2 in enumerate(nomes_cupons):                                                           # Para cada cupom n2
                df_gran.loc[n1, n2] = granularidade(cupons_tokens[n1], cupons_tokens[n2])                   # Calcula a granularidade entre n1 e n2

        # Matriz simétrica para visualização (heatmap mais intuitivo)
        df_gran_sim = pd.DataFrame(index=nomes_cupons, columns=nomes_cupons, dtype=float)                   # DataFrame para matriz simétrica
        for n1 in nomes_cupons:                                                                             # Para cada cupom n1
            for n2 in nomes_cupons:                                                                         # Para cada cupom n2
                g1 = granularidade(cupons_tokens[n1], cupons_tokens[n2])                                    # Calculla a granularidade A em B
                g2 = granularidade(cupons_tokens[n2], cupons_tokens[n1])                                    # Calculla a granularidade B em A
                df_gran_sim.loc[n1, n2] = (g1 + g2)/2                                                       # Média das granularidades para simetria        

        ws5.append([])                                                                                      # linha em branco
        ws5.append(["Granularidade (%)"] + list(df_gran.columns))                                           # Cabeçaalho da matriz de granularidade 
        for idx, row in df_gran.iterrows():                                                                 # Para cada linha da matriz de granularidade 
            ws5.append([idx] + row.tolist())                                                                # Adiciona a linha na planilha

        print("\n" + "="*70)
        print("GERANDO VISUALIZAÇÕES MELHORADAS")
        print("="*70)
        
        # -----------------------------
        # GRÁFICO 1: MATRIZ DE GRANULARIDADE
        # -----------------------------
        fig1, ax1 = plt.subplots(figsize=(max(len(nomes_cupons)*1.2, 10), max(len(nomes_cupons)*1.0, 8)))
        
        # Criar heatmap
        sns.heatmap(df_gran_sim, annot=True, fmt=".1f", cmap="RdYlGn", 
                    cbar_kws={'label':'Granularidade (%)'}, ax=ax1, 
                    vmin=0, vmax=100, linewidths=0.5, linecolor='gray')
        
        ax1.set_title(f"Matriz de Granularidade Entre Modelos OCR", 
                     fontsize=16, fontweight='bold', pad=20)
        ax1.set_xlabel("Modelo Comparado", fontsize=12, fontweight='bold')
        ax1.set_ylabel("Modelo Base", fontsize=12, fontweight='bold')
        
        # Adicionar texto explicativo
        explicacao = ("Granularidade: Percentual de tokens entre os modelos \n" +
                     "Valores mais altos (verde) = maior similaridade | Valores mais baixos (vermelho) = menor similaridade")
        fig1.text(0.5, 0.02, explicacao, ha='center', fontsize=10, 
                 style='italic', wrap=True, bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.3))
        
        plt.tight_layout(rect=[0, 0.05, 1, 1])
        png_granularidade = os.path.join(output_dir, f"01_matriz_granularidade_{base_nome}.png")
        plt.savefig(png_granularidade, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"\n[1/4] Matriz de granularidade salva: {png_granularidade}")
        
        # -----------------------------
        # GRÁFICO 2: ACURÁCIA DOS MODELOS
        # -----------------------------
        fig2, ax2 = plt.subplots(figsize=(max(len(nomes_cupons)*0.8, 10), 8))
        
        # Extrair métricas de cada modelo
        modelos = []
        acuracias = []
        precisoes = []
        recalls = []
        f1s = []
        
        for nome in nomes_cupons:
            modelos.append(nome)
            acuracias.append(metricas_total[nome]["Acuracia"] * 100)
            precisoes.append(metricas_total[nome]["Precisao"] * 100)
            recalls.append(metricas_total[nome]["Recall"] * 100)
            f1s.append(metricas_total[nome]["F1"] * 100)
        
        # Criar gráfico de barras agrupadas
        x = range(len(modelos))
        width = 0.2
        
        bars1 = ax2.bar([i - 1.5*width for i in x], acuracias, width, label='Acurácia', color='#2E86AB')
        #bars2 = ax2.bar([i - 0.5*width for i in x], precisoes, width, label='Precisão', color='#A23B72')
        #bars3 = ax2.bar([i + 0.5*width for i in x], recalls, width, label='Recall', color='#F18F01')
        #bars4 = ax2.bar([i + 1.5*width for i in x], f1s, width, label='F1-Score', color='#C73E1D')
        
        # Adicionar valores nas barras
        for bars in [bars1]:
            for bar in bars:
                height = bar.get_height()
                ax2.text(bar.get_x() + bar.get_width()/2., height,
                        f'{height:.1f}%', ha='center', va='bottom', fontsize=8)
        
        ax2.set_xlabel('Modelos OCR', fontsize=12, fontweight='bold')
        ax2.set_ylabel('Percentual (%)', fontsize=12, fontweight='bold')
        ax2.set_title(f'Desempenho dos Modelos OCR\nBase de Referência: {base_nome}', 
                     fontsize=16, fontweight='bold', pad=20)
        ax2.set_xticks(x)
        ax2.set_xticklabels(modelos, rotation=45, ha='right')
        ax2.legend(loc='upper left', fontsize=10)
        ax2.set_ylim(0, 110)
        ax2.grid(axis='y', linestyle='--', alpha=0.3)
        
        # Adicionar linha de referência em 100%
        ax2.axhline(y=100, color='green', linestyle='--', linewidth=1, alpha=0.5, label='Ideal (100%)')
        
        # Adicionar explicação das métricas
        explicacao_metricas = (
            "Acurácia: % de itens corretamente identificados" # | Precisão: % de itens identificados que são corretos\n" +
            #"Recall: % de itens corretos que foram identificados | F1-Score: média harmônica entre Precisão e Recall"
        )
        fig2.text(0.5, 0.02, explicacao_metricas, ha='center', fontsize=9, 
                 style='italic', wrap=True, bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.3))
        
        plt.tight_layout(rect=[0, 0.08, 1, 1])
        png_acuracia = os.path.join(output_dir, f"02_metricas_desempenho_{base_nome}.png")
        plt.savefig(png_acuracia, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"[2/4] Gráfico de métricas salvo: {png_acuracia}")

        # -----------------------------
        # GRÁFICO 3: GRANULARIDADE BASE vs MODELOS (NOVO!)
        # -----------------------------
        fig3, ax3 = plt.subplots(figsize=(max(len(nomes_cupons)*0.8, 10), 8))
        
        # Calcular granularidade da base em relação a cada modelo
        gran_base_para_modelos = []
        gran_modelos_para_base = []
        
        for nome in nomes_cupons:
            # Granularidade: quanto da BASE está presente no MODELO
            gran_b_m = granularidade(base, cupons_tokens[nome])
            gran_base_para_modelos.append(gran_b_m)
            
            # Granularidade: quanto do MODELO está presente na BASE
            gran_m_b = granularidade(cupons_tokens[nome], base)
            gran_modelos_para_base.append(gran_m_b)
        
        # Criar gráfico de barras agrupadas
        x = range(len(modelos))
        width = 0.35
        
        bars1 = ax3.bar([i - width/2 for i in x], gran_base_para_modelos, width, 
                       label='Base → Modelo (Recall / Cobertura)', color='#4ECDC4')
        bars2 = ax3.bar([i + width/2 for i in x], gran_modelos_para_base, width, 
                       label='Modelo → Base (Pureza / Aderência)', color='#FF6B6B')
        
        # Adicionar valores nas barras
        for bars in [bars1, bars2]:
            for bar in bars:
                height = bar.get_height()
                ax3.text(bar.get_x() + bar.get_width()/2., height,
                        f'{height:.1f}%', ha='center', va='bottom', fontsize=9, fontweight='bold')
        
        ax3.set_xlabel('Modelos OCR', fontsize=12, fontweight='bold')
        ax3.set_ylabel('Granularidade (%)', fontsize=12, fontweight='bold')
        ax3.set_title(f'Granularidade de Tokens: Base de Referência vs Modelos OCR\nBase de Referência: {base_nome}',
                     fontsize=16, fontweight='bold', pad=20)
        ax3.set_xticks(x)
        ax3.set_xticklabels(modelos, rotation=45, ha='right')
        ax3.legend(loc='upper left', fontsize=10)
        ax3.set_ylim(0, 110)
        ax3.grid(axis='y', linestyle='--', alpha=0.3)
        
        # Linha de referência em 100%
        ax3.axhline(y=100, color='green', linestyle='--', linewidth=1, alpha=0.5, label='Ideal (100%)')
        
        # Explicação
        explicacao_gran = (
            #"Base → Modelo: % dos tokens da referência que foram capturados pelo modelo (cobertura)\n" +
            #"Modelo → Base: % dos tokens do modelo que estão corretos segundo a referência (precisão)"
            "Base → Modelo (Recall): porcentagem dos tokens da base de referência"
            "que foram encontrados no modelo OCR (cobertura da informação).\n"
            "Modelo → Base (Pureza): porcentagem dos tokens gerados pelo modelo OCR "
            "que também existem na base de referência (aderência lexical)."
        )
        fig3.text(0.5, 0.02, explicacao_gran, ha='center', fontsize=9, 
                 style='italic', wrap=True, bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.3))
        
        plt.tight_layout(rect=[0, 0.08, 1, 1])
        png_gran_base = os.path.join(output_dir, f"03_granularidade_base_vs_modelos_{base_nome}.png")
        plt.savefig(png_gran_base, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"[3/4] Gráfico de granularidade Base vs Modelos salvo: {png_gran_base}")

        # -----------------------------
        # Gráfico longo de acertos/erros (MANTIDO DO ORIGINAL)
        # -----------------------------
        itens_base = df_resultado["ITEM_BASE"].tolist()                             # Para o eixo X do gráfico
        x = range(len(itens_base))                                                  # Eixo X numérico com base no número de itens
        plt.figure(figsize=(max(len(itens_base)*0.15, 10), max(len(nomes_cupons)*0.5, 6)))

        # Preparar dados para heatmap
        heatmap_data = df_resultado[nomes_cupons].T  # Transpor para cupons nas linhas

        # Criar heatmap
        sns.heatmap(heatmap_data, 
                    annot=True, 
                    fmt='d',
                    cmap=['#ff6b6b', '#51cf66'],  # Vermelho para 0, Verde para 1
                    cbar=False,
                    linewidths=0.5,
                    linecolor='gray')

        plt.title(f"Acertos/Erros por Cupom e Item ({len(itens_base)} itens, {len(nomes_cupons)} cupons)")
        plt.xlabel("Item Base")
        plt.ylabel("Cupom")
        plt.xticks(ticks=np.arange(len(itens_base)) + 0.5, 
                   labels=itens_base, 
                   rotation=90, 
                   fontsize=6)
        plt.yticks(rotation=0)

        plt.tight_layout()
        save_name = os.path.join(output_dir, f"04_heatmap_comparacao_{base_nome}.png")
        plt.savefig(save_name, dpi=150)
        plt.close()
        print(f"[4/4] Gráfico de comparação salvo: {save_name}")                    # Mensagem de status
        
        print("\n" + "="*70)
        print("✓ Todas as visualizações foram geradas com sucesso!")
        print("="*70)
    else:                                                                           # Se não há dados avisa
        print("AVISO: Não há dados suficientes para gerar o gráfico")               # Mensagem de aviso
               
    output_file = os.path.join(output_dir, f"comparacao_{base_nome}.xlsx")          # Salvar arquivo XLSX
    wb.save(output_file)                                                            # Salva o arquivo XLSX
    print(f"\nArquivo Excel gerado: {output_file}")                                 # Mensagem de status
    
    if len(base) == 0:                                                                                                  # Verifica se a base está vazia
        print("\nAVISO: A base não contém tokens! Verifique o formato do arquivo.")                                     # Se estiver, avisa o usuário
        print("Para JSON, certifique-se que possui a estrutura: {\"itens\": [{\"codigo\": ..., \"descricao\": ...}]}")  # Se estiver, avisa o usuário#
        print("Para TXT, certifique-se que possui linhas no formato: OCR='...'")                                        # Se estiver, avisa o usuário

if __name__ == "__main__":  # Execução do main
    main()                  # Chama a função principal